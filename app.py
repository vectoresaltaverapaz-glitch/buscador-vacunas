import logging
import os
from flask import Flask, render_template_string, request, jsonify, Response
from flask_httpauth import HTTPBasicAuth
from datetime import datetime, timedelta
from io import BytesIO
import libsql_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
auth = HTTPBasicAuth()

USUARIOS = {
    "admin": "vacunas2025",
    "consulta": "buscar123"
}

@auth.verify_password
def verify_password(usuario, password):
    return usuario in USUARIOS and USUARIOS[usuario] == password

def get_turso_client():
    url = os.getenv("TURSO_URL")
    token = os.getenv("TURSO_TOKEN")
    if not url or not token:
        logger.error("❌ Faltan variables de entorno TURSO_URL o TURSO_TOKEN")
        raise Exception("Faltan variables de entorno")
    logger.info(f"✅ Conectando a Turso: {url}")
    return libsql_client.create_client_sync(url=url, auth_token=token)

# ==================================================
# COLUMNAS REALES (lista fija - todas las columnas de la BD)
# ==================================================

COLUMNAS_REALES = [
    "rowid", "año", "no.", "área", "distrito", "servicio",
    "departamento", "municipio", "cui_nino", "nombre_nino",
    "día", "mes", "año_1", "departamento.1", "municipio.1",
    "comunidad", "hombre", "mujer", "pueblo", "comunidad.1",
    "cui_responsable", "nombre_responsable", "día.1", "mes.1",
    "año.1", "departamento.2", "municipio.2", "comunidad.2",
    "calle,_avenida,_zona,_lote,", "telefono", "falleció",
    "hep._b", "bcg", "1a._(fipv)", "2a._(fipv)", "1a._(ipv)",
    "1a._(historico)", "2a._(opv)", "3a._(opv)", "1a.", "2a.",
    "3a.", "1a..1", "2a..1", "1a..2", "2a..2", "spr_1",
    "neumo-_r1", "spr_2", "r1_(opv)", "r1_(dpt)", "r2_(opv)",
    "r2_(dpt)", "1a..3", "2a..3", "1a..4", "2a..4", "1a..5",
    "2a..5", "1a._(fipv).1", "2a._(fipv).1", "1a._(ipv).1",
    "1a._(historico).1", "2a._(opv).1", "3a._(opv).1",
    "r1_(opv).1", "r2_(opv).1", "1a..6", "2a..6", "3a..1",
    "r1", "r2", "spr_1.1", "spr_2.1", "1a..7", "2a..7", "3a..2"
]

# ==================================================
# RENOMBRES (títulos bonitos para mostrar en pantalla)
# ==================================================

RENOMBRES = {
    "año": "Año",
    "distrito": "Distrito",
    "servicio": "Servicio Salud",
    "departamento": "Departamento",
    "municipio": "Municipio",
    "cui_nino": "CUI Niño",
    "nombre_nino": "Nombre Niño",
    "cui_responsable": "CUI Responsable",
    "nombre_responsable": "Nombre Responsable",
    "telefono": "Teléfono",
    "falleció": "Falleció",
    "genero": "Género",
    "fecha_nac_nino": "Fecha Nac. Niño",
    "fecha_nac_responsable": "Fecha Nac. Responsable",
    # Vacunas
    "hep._b": "Hepatitis B (<1 año)",
    "bcg": "BCG (<1 año)",
    "1a._(fipv)": "Polio 1 (<1 año)",
    "2a._(fipv)": "Polio 2 (<1 año)",
    "1a._(ipv)": "Polio IPV (<1 año)",
    "1a._(historico)": "Polio Histórico (<1 año)",
    "2a._(opv)": "OPV 2 (<1 año)",
    "3a._(opv)": "OPV 3 (<1 año)",
    "1a.": "Pentavalente 1 (<1 año)",
    "2a.": "Pentavalente 2 (<1 año)",
    "3a.": "Pentavalente 3 (<1 año)",
    "1a..1": "Rotavirus 1 (<1 año)",
    "2a..1": "Rotavirus 2 (<1 año)",
    "1a..2": "Neumococo 1 (<1 año)",
    "2a..2": "Neumococo 2 (<1 año)",
    "spr_1": "SPR 1 (12 meses)",
    "neumo-_r1": "Neumococo R1 (12 meses)",
    "spr_2": "SPR 2 (18 meses)",
    "r1_(opv)": "Refuerzo OPV (18 meses)",
    "r1_(dpt)": "Refuerzo DPT (18 meses)",
    "r2_(opv)": "Refuerzo OPV (4-7 años)",
    "r2_(dpt)": "Refuerzo DPT (4-7 años)",
    "1a..3": "Influenza 1 (6-11 meses)",
    "2a..3": "Influenza 2 (6-11 meses)",
    "1a..4": "Influenza 1 (12-23 meses)",
    "2a..4": "Influenza 2 (12-23 meses)",
    "1a..5": "Influenza 1 (24-35 meses)",
    "2a..5": "Influenza 2 (24-35 meses)",
    "1a._(fipv).1": "Polio fIPV 1 (1-7 años)",
    "2a._(fipv).1": "Polio fIPV 2 (1-7 años)",
    "1a._(ipv).1": "Polio IPV Refuerzo (1-7 años)",
    "1a._(historico).1": "Polio Histórico Refuerzo (1-7 años)",
    "2a._(opv).1": "OPV 2 Refuerzo (1-7 años)",
    "3a._(opv).1": "OPV 3 Refuerzo (1-7 años)",
    "r1_(opv).1": "Refuerzo OPV 1 (1-7 años)",
    "r2_(opv).1": "Refuerzo OPV 2 (1-7 años)",
    "1a..6": "Pentavalente Ref 1 (1-7 años)",
    "2a..6": "Pentavalente Ref 2 (1-7 años)",
    "3a..1": "Pentavalente Ref 3 (1-7 años)",
    "r1": "Refuerzo 1 (1-7 años)",
    "r2": "Refuerzo 2 (1-7 años)",
    "spr_1.1": "SPR Refuerzo 1 (1-7 años)",
    "spr_2.1": "SPR Refuerzo 2 (1-7 años)",
    "1a..7": "Otras Vacunas 1 (1-7 años)",
    "2a..7": "Otras Vacunas 2 (1-7 años)",
    "3a..2": "Otras Vacunas 3 (1-7 años)"
}

# ==================================================
# COLUMNAS FECHA (para convertir números a fecha)
# ==================================================

COLUMNAS_FECHA = {
    "hep._b", "bcg",
    "1a._(fipv)", "2a._(fipv)", "1a._(ipv)",
    "1a._(historico)", "2a._(opv)", "3a._(opv)",
    "1a.", "2a.", "3a.",
    "1a..1", "2a..1",
    "1a..2", "2a..2",
    "spr_1", "neumo-_r1", "spr_2",
    "r1_(opv)", "r1_(dpt)",
    "r2_(opv)", "r2_(dpt)",
    "1a..3", "2a..3",
    "1a..4", "2a..4",
    "1a..5", "2a..5",
    "1a._(fipv).1", "2a._(fipv).1", "1a._(ipv).1",
    "1a._(historico).1",
    "2a._(opv).1", "3a._(opv).1",
    "r1_(opv).1", "r2_(opv).1",
    "1a..6", "2a..6", "3a..1",
    "r1", "r2",
    "spr_1.1", "spr_2.1",
    "1a..7", "2a..7", "3a..2"
}

# ==================================================
# COLUMNAS EXCLUIDAS (no se muestran en la tabla)
# ==================================================

COLUMNAS_EXCLUIDAS = [
    "rowid", "no.", "área", "pueblo", "comunidad",
    "departamento.1", "municipio.1", "comunidad.1",
    "comunidad.2", "calle,_avenida,_zona,_lote,",
    "día", "mes", "año_1", "día.1", "mes.1", "año.1",
    "hombre", "mujer"
]

# ==================================================
# FUNCIONES DE PROCESAMIENTO
# ==================================================

def limpiar_numero(valor):
    try:
        if valor is None:
            return ""
        numero = float(valor)
        if numero.is_integer():
            return str(int(numero))
        return str(numero)
    except:
        return str(valor)

def convertir_fecha_excel(valor):
    try:
        if valor is None or valor == "":
            return ""
        numero = float(valor)
        if 20000 <= numero <= 60000:
            fecha_base = datetime(1899, 12, 30)
            fecha = fecha_base + timedelta(days=int(numero))
            return fecha.strftime("%d/%m/%Y")
        return limpiar_numero(valor)
    except:
        return str(valor)

def procesar_valor(columna, valor):
    if columna in COLUMNAS_FECHA:
        return convertir_fecha_excel(valor)
    return limpiar_numero(valor)

def obtener_genero(row):
    hombre = row.get("hombre")
    mujer = row.get("mujer")
    if hombre == 'X':
        return "Hombre"
    elif mujer == 'X':
        return "Mujer"
    else:
        return "No especificado"

def formatear_fecha(dia, mes, año):
    """Formatea día, mes, año a DD/MM/YYYY si son válidos."""
    try:
        dia = int(dia) if dia else None
        mes = int(mes) if mes else None
        año = int(año) if año else None
        if dia and mes and año and 1 <= dia <= 31 and 1 <= mes <= 12 and 1900 <= año <= 2100:
            return f"{dia:02d}/{mes:02d}/{año}"
        return ""
    except:
        return ""

# ==================================================
# HTML
# ==================================================

HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Buscador de Vacunas</title>
<style>
body{ font-family:Arial; background:#eef2f7; padding:20px; }
.container{ background:white; padding:20px; border-radius:10px; }
input,select,button{ padding:10px; margin:5px; }
.filtros{ margin-bottom:15px; }
.filtros select, .filtros input{ margin-bottom:5px; }
.botones{ display:flex; gap:10px; flex-wrap:wrap; align-items:center; }
table{ width:100%; border-collapse:collapse; margin-top:20px; font-size:12px; }
th,td{ border:1px solid #ccc; padding:5px; }
th{ background:#1e466e; color:white; position:sticky; top:0; }

.spinner {
    display: none;
    border: 6px solid #f3f3f3;
    border-top: 6px solid #1e466e;
    border-radius: 50%;
    width: 40px;
    height: 40px;
    margin: 20px auto;
    animation: spin 0.8s linear infinite;
}
@keyframes spin {
    0% { transform: rotate(0deg); }
    100% { transform: rotate(360deg); }
}
</style>
</head>
<body>
<div class="container">
<h2>🔍 Buscador de Vacunas</h2>

<div class="filtros">
    <div>
        <select id="distrito">
            <option value="">TODOS LOS DISTRITOS</option>
            <option value="COBAN">COBAN</option>
            <option value="SAN PEDRO CARCHA">SAN PEDRO CARCHA</option>
            <option value="TACTIC">TACTIC</option>
            <option value="LANQUIN">LANQUIN</option>
            <option value="CHISEC">CHISEC</option>
            <option value="CAHABON">CAHABON</option>
            <option value="SENAHU">SENAHU</option>
            <option value="FRAY BARTOLOME">FRAY BARTOLOME</option>
            <option value="PANZÓS">PANZÓS</option>
            <option value="SAN JUAN CHAMELCO">SAN JUAN CHAMELCO</option>
            <option value="TUCURÚ">TUCURÚ</option>
            <option value="TAMAHÚ">TAMAHÚ</option>
            <option value="SANTA CRUZ VERAPAZ">SANTA CRUZ VERAPAZ</option>
            <option value="CAMPUR">CAMPUR</option>
            <option value="COBÁN">COBÁN</option>
            <option value="RAXRUHA">RAXRUHA</option>
            <option value="CHAHAL">CHAHAL</option>
            <option value="LA TINTA">LA TINTA</option>
            <option value="TELEMÁN">TELEMÁN</option>
        </select>

        <select id="tipo">
            <option value="nombre_nino">Nombre Niño</option>
            <option value="nombre_responsable">Nombre Responsable</option>
            <option value="cui_nino">CUI Niño</option>
            <option value="cui_responsable">CUI Responsable</option>
        </select>

        <select id="genero">
            <option value="">Género (todos)</option>
            <option value="Hombre">Hombre</option>
            <option value="Mujer">Mujer</option>
            <option value="No especificado">No especificado</option>
        </select>

        <input type="text" id="search" placeholder="Buscar..." style="width:200px;">
    </div>

    <div style="margin-top:10px;">
        <label>Fecha Nacimiento Niño:</label>
        <input type="number" id="dia_nac" placeholder="Día" style="width:60px;">
        <input type="number" id="mes_nac" placeholder="Mes" style="width:60px;">
        <input type="number" id="ano_nac" placeholder="Año" style="width:80px;">

        <label style="margin-left:15px;">Fecha Nacimiento Responsable:</label>
        <input type="number" id="dia_madre" placeholder="Día" style="width:60px;">
        <input type="number" id="mes_madre" placeholder="Mes" style="width:60px;">
        <input type="number" id="ano_madre" placeholder="Año" style="width:80px;">
    </div>

    <div class="botones" style="margin-top:10px;">
        <button id="btnBuscar" onclick="iniciarBusqueda()">🔍 Buscar</button>
        <button id="btnCancelar" onclick="cancelarBusqueda()" style="display:none; background:#f39c12; color:white; border:none;">⏹ Cancelar</button>
        <button onclick="exportarExcel()">📊 Exportar Excel</button>
        <button onclick="limpiar()" style="background:#e74c3c; color:white; border:none;">🔄 Nueva búsqueda</button>
        <span style="margin-left:15px;font-size:14px;color:#555;">Mostrando hasta 30 registros</span>
    </div>
</div>

<div id="resultado"></div>
<div class="spinner" id="spinner"></div>
</div>

<script>
var abortController = null;

function iniciarBusqueda(){
    if (abortController) {
        abortController.abort();
    }

    document.getElementById('spinner').style.display = 'block';
    document.getElementById('btnBuscar').disabled = true;
    document.getElementById('btnCancelar').style.display = 'inline-block';
    document.getElementById('resultado').innerHTML = '';

    abortController = new AbortController();

    let q = document.getElementById('search').value;
    let tipo = document.getElementById('tipo').value;
    let distrito = document.getElementById('distrito').value;
    let genero = document.getElementById('genero').value;
    let dia_nac = document.getElementById('dia_nac').value;
    let mes_nac = document.getElementById('mes_nac').value;
    let ano_nac = document.getElementById('ano_nac').value;
    let dia_madre = document.getElementById('dia_madre').value;
    let mes_madre = document.getElementById('mes_madre').value;
    let ano_madre = document.getElementById('ano_madre').value;

    let url = `/buscar?q=${encodeURIComponent(q)}&tipo=${tipo}&distrito=${encodeURIComponent(distrito)}&genero=${encodeURIComponent(genero)}`;
    if (dia_nac) url += `&dia_nac=${dia_nac}`;
    if (mes_nac) url += `&mes_nac=${mes_nac}`;
    if (ano_nac) url += `&ano_nac=${ano_nac}`;
    if (dia_madre) url += `&dia_madre=${dia_madre}`;
    if (mes_madre) url += `&mes_madre=${mes_madre}`;
    if (ano_madre) url += `&ano_madre=${ano_madre}`;

    fetch(url, { signal: abortController.signal })
    .then(r => r.json())
    .then(data => {
        finalizarBusqueda();
        if (data.error) {
            document.getElementById('resultado').innerHTML = `<p style="color:red;">❌ Error: ${data.error}</p>`;
            return;
        }
        if (!data.columnas || !data.rows) {
            document.getElementById('resultado').innerHTML = '<p>⚠️ No se encontraron resultados.</p>';
            return;
        }
        let html = '';
        html += `<p>Total resultados: ${data.total}</p>`;
        html += '<div style="overflow:auto; max-height:700px;">';
        html += '<table><tr>';
        data.columnas.forEach(c => { html += `<th>${c}</th>`; });
        html += '</tr>';
        data.rows.forEach(r => {
            html += '<tr>';
            r.forEach(c => { html += `<td>${c ?? ''}</td>`; });
            html += '</tr>';
        });
        html += '</table></div>';
        document.getElementById('resultado').innerHTML = html;
    })
    .catch(err => {
        if (err.name === 'AbortError') {
            document.getElementById('resultado').innerHTML = '<p>⏹ Búsqueda cancelada.</p>';
        } else {
            finalizarBusqueda();
            document.getElementById('resultado').innerHTML = `<p style="color:red;">❌ Error: ${err}</p>`;
        }
    })
    .finally(() => {
        finalizarBusqueda();
        abortController = null;
    });
}

function cancelarBusqueda(){
    if (abortController) {
        abortController.abort();
    }
}

function finalizarBusqueda(){
    document.getElementById('spinner').style.display = 'none';
    document.getElementById('btnBuscar').disabled = false;
    document.getElementById('btnCancelar').style.display = 'none';
}

function limpiar(){
    if (abortController) {
        abortController.abort();
        abortController = null;
    }
    document.getElementById('search').value = '';
    document.getElementById('tipo').value = 'nombre_nino';
    document.getElementById('distrito').value = '';
    document.getElementById('genero').value = '';
    document.getElementById('dia_nac').value = '';
    document.getElementById('mes_nac').value = '';
    document.getElementById('ano_nac').value = '';
    document.getElementById('dia_madre').value = '';
    document.getElementById('mes_madre').value = '';
    document.getElementById('ano_madre').value = '';
    document.getElementById('resultado').innerHTML = '';
    finalizarBusqueda();
    iniciarBusqueda();
}

function exportarExcel(){
    let q = document.getElementById('search').value;
    let tipo = document.getElementById('tipo').value;
    let distrito = document.getElementById('distrito').value;
    let genero = document.getElementById('genero').value;
    let dia_nac = document.getElementById('dia_nac').value;
    let mes_nac = document.getElementById('mes_nac').value;
    let ano_nac = document.getElementById('ano_nac').value;
    let dia_madre = document.getElementById('dia_madre').value;
    let mes_madre = document.getElementById('mes_madre').value;
    let ano_madre = document.getElementById('ano_madre').value;

    let url = `/exportar?q=${encodeURIComponent(q)}&tipo=${tipo}&distrito=${encodeURIComponent(distrito)}&genero=${encodeURIComponent(genero)}`;
    if (dia_nac) url += `&dia_nac=${dia_nac}`;
    if (mes_nac) url += `&mes_nac=${mes_nac}`;
    if (ano_nac) url += `&ano_nac=${ano_nac}`;
    if (dia_madre) url += `&dia_madre=${dia_madre}`;
    if (mes_madre) url += `&mes_madre=${mes_madre}`;
    if (ano_madre) url += `&ano_madre=${ano_madre}`;

    window.location.href = url;
}
</script>
</body>
</html>
"""

# ==================================================
# RUTAS
# ==================================================

@app.route('/')
@auth.login_required
def index():
    return render_template_string(HTML)

# ==================================================
# BUSCAR CON FTS (Full-Text Search) y límite dinámico
# ==================================================

@app.route('/buscar')
@auth.login_required
def buscar():
    query = request.args.get('q', '').strip()
    tipo = request.args.get('tipo', 'nombre_nino')
    distrito = request.args.get('distrito', '').strip()
    genero = request.args.get('genero', '').strip()
    dia_nac = request.args.get('dia_nac', '').strip()
    mes_nac = request.args.get('mes_nac', '').strip()
    ano_nac = request.args.get('ano_nac', '').strip()
    dia_madre = request.args.get('dia_madre', '').strip()
    mes_madre = request.args.get('mes_madre', '').strip()
    ano_madre = request.args.get('ano_madre', '').strip()
    
    logger.info(f"🔍 Búsqueda: query='{query}', tipo='{tipo}', distrito='{distrito}', genero='{genero}'")
    
    try:
        client = get_turso_client()
        
        # --- 1. FTS para búsqueda de texto ---
        fts_condition = ""
        fts_params = []
        
        if query:
            if tipo == "nombre_nino":
                fts_query = f"{query}*"
                fts_sql = f"""
                    SELECT rowid FROM fts_data 
                    WHERE nombre MATCH ? 
                    ORDER BY rank
                    LIMIT 100
                """
                fts_result = client.execute(fts_sql, [fts_query])
                if hasattr(fts_result, 'rows') and callable(fts_result.rows):
                    fts_rows = fts_result.rows()
                else:
                    fts_rows = list(fts_result)
                
                if fts_rows:
                    rowids = [str(row[0]) for row in fts_rows]
                    fts_condition = f'"rowid" IN ({",".join(rowids)})'
                    logger.info(f"🔍 FTS encontró {len(rowids)} filas")
                else:
                    client.close()
                    return jsonify({"rows": [], "columnas": [], "total": 0})
            
            elif tipo == "nombre_responsable":
                fts_condition = f'"nombre_responsable" LIKE ?'
                fts_params.append(f"%{query}%")
        
        # --- 2. Construir condiciones adicionales ---
        condiciones = []
        parametros = []
        
        if query and not fts_condition:
            if tipo == "nombre_nino":
                condiciones.append('"nombre_nino" LIKE ?')
                parametros.append(f"%{query}%")
            elif tipo == "nombre_responsable":
                condiciones.append('"nombre_responsable" LIKE ?')
                parametros.append(f"%{query}%")
            elif tipo == "cui_nino":
                condiciones.append('"cui_nino" = ?')
                parametros.append(query)
            elif tipo == "cui_responsable":
                condiciones.append('"cui_responsable" = ?')
                parametros.append(query)
        
        if distrito:
            condiciones.append('UPPER("distrito") LIKE ?')
            parametros.append(f"%{distrito.upper()}%")
        
        if genero:
            if genero == "Hombre":
                condiciones.append('"hombre" = "X"')
            elif genero == "Mujer":
                condiciones.append('"mujer" = "X"')
            elif genero == "No especificado":
                condiciones.append('"hombre" != "X" AND "mujer" != "X"')
        
        if dia_nac:
            condiciones.append('"día" = ?')
            parametros.append(dia_nac)
        if mes_nac:
            condiciones.append('"mes" = ?')
            parametros.append(mes_nac)
        if ano_nac:
            condiciones.append('"año_1" = ?')
            parametros.append(ano_nac)
        
        if dia_madre:
            condiciones.append('"día.1" = ?')
            parametros.append(dia_madre)
        if mes_madre:
            condiciones.append('"mes.1" = ?')
            parametros.append(mes_madre)
        if ano_madre:
            condiciones.append('"año.1" = ?')
            parametros.append(ano_madre)
        
        where_parts = []
        if fts_condition:
            where_parts.append(fts_condition)
        where_parts.extend(condiciones)
        
        where = " AND ".join(where_parts) if where_parts else ""
        all_params = fts_params + parametros
        
        # --- 3. Límite dinámico según cantidad de filtros ---
        num_filtros = len(condiciones)
        if num_filtros > 2:
            limite = 15
        else:
            limite = 30
        
        sql = f"""
        SELECT * FROM datos_completos
        {f'WHERE {where}' if where else ''}
        LIMIT {limite}
        """
        logger.info(f"📝 SQL: {sql}")
        logger.info(f"📦 Parámetros: {all_params}")
        logger.info(f"📌 Límite aplicado: {limite}")
        
        result = client.execute(sql, all_params)
        if hasattr(result, 'rows') and callable(result.rows):
            rows = result.rows()
        else:
            rows = list(result)
        
        logger.info(f"📊 Filas obtenidas: {len(rows)}")
        
        if not rows:
            client.close()
            return jsonify({"rows": [], "columnas": [], "total": 0})
        
        rows_dict = [dict(zip(COLUMNAS_REALES, row)) for row in rows]
        
        # Calcular género y fechas
        for row in rows_dict:
            row["genero"] = obtener_genero(row)
            row["fecha_nac_nino"] = formatear_fecha(row.get("día"), row.get("mes"), row.get("año_1"))
            row["fecha_nac_responsable"] = formatear_fecha(row.get("día.1"), row.get("mes.1"), row.get("año.1"))
        
        columnas_finales = [c for c in COLUMNAS_REALES if c not in COLUMNAS_EXCLUIDAS]
        try:
            idx = columnas_finales.index('nombre_responsable')
            columnas_finales.insert(idx+1, 'fecha_nac_responsable')
        except ValueError:
            columnas_finales.append('fecha_nac_responsable')
        try:
            idx = columnas_finales.index('nombre_nino')
            columnas_finales.insert(idx+1, 'fecha_nac_nino')
        except ValueError:
            columnas_finales.append('fecha_nac_nino')
        
        titulos_columnas = [RENOMBRES.get(c, c) for c in columnas_finales]
        
        resultados = []
        for row in rows_dict:
            fila = []
            for c in columnas_finales:
                if c in ['fecha_nac_nino', 'fecha_nac_responsable', 'genero']:
                    fila.append(row.get(c, ''))
                else:
                    fila.append(procesar_valor(c, row.get(c)))
            resultados.append(fila)
        
        client.close()
        logger.info(f"✅ Resultados devueltos: {len(resultados)}")
        return jsonify({
            "rows": resultados,
            "columnas": titulos_columnas,
            "total": len(resultados)
        })
    
    except Exception as e:
        logger.error(f"❌ Error en /buscar: {str(e)}", exc_info=True)
        return jsonify({
            "rows": [],
            "columnas": [],
            "total": 0,
            "error": str(e)
        })

# ==================================================
# EXPORTAR CON LÍMITE DE 30 REGISTROS
# ==================================================

MAX_EXPORT_ROWS = 30

@app.route('/exportar')
@auth.login_required
def exportar():
    query = request.args.get('q', '').strip()
    tipo = request.args.get('tipo', 'nombre_nino')
    distrito = request.args.get('distrito', '').strip()
    genero = request.args.get('genero', '').strip()
    dia_nac = request.args.get('dia_nac', '').strip()
    mes_nac = request.args.get('mes_nac', '').strip()
    ano_nac = request.args.get('ano_nac', '').strip()
    dia_madre = request.args.get('dia_madre', '').strip()
    mes_madre = request.args.get('mes_madre', '').strip()
    ano_madre = request.args.get('ano_madre', '').strip()
    
    try:
        client = get_turso_client()
        
        condiciones = []
        parametros = []
        
        if query:
            if tipo == "nombre_nino":
                condiciones.append('"nombre_nino" LIKE ?')
                parametros.append(f"%{query}%")
            elif tipo == "nombre_responsable":
                condiciones.append('"nombre_responsable" LIKE ?')
                parametros.append(f"%{query}%")
            elif tipo == "cui_nino":
                condiciones.append('"cui_nino" = ?')
                parametros.append(query)
            elif tipo == "cui_responsable":
                condiciones.append('"cui_responsable" = ?')
                parametros.append(query)
        
        if distrito:
            condiciones.append('UPPER("distrito") LIKE ?')
            parametros.append(f"%{distrito.upper()}%")
        
        if genero:
            if genero == "Hombre":
                condiciones.append('"hombre" = "X"')
            elif genero == "Mujer":
                condiciones.append('"mujer" = "X"')
            elif genero == "No especificado":
                condiciones.append('"hombre" != "X" AND "mujer" != "X"')
        
        if dia_nac:
            condiciones.append('"día" = ?')
            parametros.append(dia_nac)
        if mes_nac:
            condiciones.append('"mes" = ?')
            parametros.append(mes_nac)
        if ano_nac:
            condiciones.append('"año_1" = ?')
            parametros.append(ano_nac)
        
        if dia_madre:
            condiciones.append('"día.1" = ?')
            parametros.append(dia_madre)
        if mes_madre:
            condiciones.append('"mes.1" = ?')
            parametros.append(mes_madre)
        if ano_madre:
            condiciones.append('"año.1" = ?')
            parametros.append(ano_madre)
        
        where = " AND ".join(condiciones) if condiciones else ""
        sql = f"""
        SELECT * FROM datos_completos
        {f'WHERE {where}' if where else ''}
        LIMIT {MAX_EXPORT_ROWS}
        """
        logger.info(f"📝 SQL export: {sql}")
        
        result = client.execute(sql, parametros)
        if hasattr(result, 'rows') and callable(result.rows):
            rows = result.rows()
        else:
            rows = list(result)
        
        if not rows:
            client.close()
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados"
            ws.append(["No se encontraron registros"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': 'attachment; filename=resultados.xlsx'}
            )
        
        rows_dict = [dict(zip(COLUMNAS_REALES, row)) for row in rows]
        for row in rows_dict:
            row["genero"] = obtener_genero(row)
            row["fecha_nac_nino"] = formatear_fecha(row.get("día"), row.get("mes"), row.get("año_1"))
            row["fecha_nac_responsable"] = formatear_fecha(row.get("día.1"), row.get("mes.1"), row.get("año.1"))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        
        columnas_finales = [c for c in COLUMNAS_REALES if c not in COLUMNAS_EXCLUIDAS]
        try:
            idx = columnas_finales.index('nombre_responsable')
            columnas_finales.insert(idx+1, 'fecha_nac_responsable')
        except ValueError:
            columnas_finales.append('fecha_nac_responsable')
        try:
            idx = columnas_finales.index('nombre_nino')
            columnas_finales.insert(idx+1, 'fecha_nac_nino')
        except ValueError:
            columnas_finales.append('fecha_nac_nino')
        
        titulos_columnas = [RENOMBRES.get(c, c) for c in columnas_finales]
        ws.append(titulos_columnas)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e466e", end_color="1e466e", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row in rows_dict:
            fila = []
            for c in columnas_finales:
                if c in ['fecha_nac_nino', 'fecha_nac_responsable', 'genero']:
                    fila.append(row.get(c, ''))
                else:
                    fila.append(procesar_valor(c, row.get(c)))
            ws.append(fila)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        client.close()
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=resultados.xlsx'}
        )
    
    except Exception as e:
        logger.error(f"❌ Error en /exportar: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ==================================================
# INICIO
# ==================================================

if __name__ == '__main__':
    logger.info("🚀 Servidor iniciado...")
    app.run(host='0.0.0.0', port=5000, debug=True, threaded=True)